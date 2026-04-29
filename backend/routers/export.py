"""Export router — Excel file generation."""

from __future__ import annotations

import json
import io
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook

from schemas.data import ExportRequest
from storage.profiles import get_profile

router = APIRouter()


@router.post("/excel")
async def export_excel(body: ExportRequest) -> StreamingResponse:
    """Generate an Excel workbook with profile data and optional calibration results."""
    profile = get_profile(body.profile_id)
    if profile is None:
        raise HTTPException(status_code=404, detail=f"Profile {body.profile_id} not found")

    wb = Workbook()

    # ---- Sheet 1: Model Summary ----
    ws_summary = wb.active
    ws_summary.title = "Model Summary"

    ws_summary.append(["Motor Thermal Model v2 — Profile Summary"])
    ws_summary.append([])
    ws_summary.append(["Profile Name", profile.name])
    ws_summary.append(["Profile ID", profile.id])
    ws_summary.append(["Created", str(profile.created_at)])
    ws_summary.append(["Updated", str(profile.updated_at)])
    ws_summary.append([])
    ws_summary.append(["--- Geometry ---"])
    ws_summary.append(["D_motor [mm]", profile.geometry.D_motor_mm])
    ws_summary.append(["L_motor [mm]", profile.geometry.L_motor_mm])
    ws_summary.append(["t_housing [mm]", profile.geometry.t_housing_mm])
    ws_summary.append(["m_motor [g]", profile.geometry.m_motor_g])
    ws_summary.append(["m_housing [g]", profile.geometry.m_housing_g])
    ws_summary.append(["f_copper", profile.geometry.f_copper])
    ws_summary.append(["t_mold [mm]", profile.geometry.t_mold_mm])
    ws_summary.append([])
    ws_summary.append(["--- Material Properties ---"])
    ws_summary.append(["c_p_Cu [J/(kg*K)]", profile.material.c_p_Cu])
    ws_summary.append(["c_p_FeSi [J/(kg*K)]", profile.material.c_p_FeSi])
    ws_summary.append(["c_p_Al [J/(kg*K)]", profile.material.c_p_Al])
    ws_summary.append(["k_mold [W/(m*K)]", profile.material.k_mold])
    ws_summary.append([])
    ws_summary.append(["--- Coil Parameters ---"])
    ws_summary.append(["R0 [Ohm]", profile.coil.R0])
    ws_summary.append(["T0 [degC]", profile.coil.T0])
    ws_summary.append(["alpha [1/degC]", profile.coil.alpha])
    ws_summary.append(["n_phases", profile.coil.n_phases])

    if profile.geometry_preview is not None:
        gp = profile.geometry_preview
        ws_summary.append([])
        ws_summary.append(["--- Computed Geometry ---"])
        ws_summary.append(["C_coil [J/degC]", gp.C_coil])
        ws_summary.append(["C_core [J/degC]", gp.C_core])
        ws_summary.append(["C_housing [J/degC]", gp.C_housing])
        ws_summary.append(["A_interface [m^2]", gp.A_interface_m2])
        ws_summary.append(["A_housing [m^2]", gp.A_housing_m2])
        ws_summary.append(["R2_mold_init [degC/W]", gp.R2_mold_init])
        ws_summary.append(["R3_nat_init [degC/W]", gp.R3_nat_init])
        ws_summary.append(["tau_coil [s]", gp.tau_coil_s])

    # ---- Sheet 2: Calibration Results (if available) ----
    # Load calib_result from raw profile JSON
    profiles_dir = Path.home() / ".mtm_v2" / "profiles"
    fp = profiles_dir / f"{profile.id}.json"
    calib_data: Optional[dict] = None
    if fp.exists():
        raw = json.loads(fp.read_text(encoding="utf-8"))
        calib_data = raw.get("calib_result")

    if calib_data is not None:
        ws_calib = wb.create_sheet("Calibration Detail")
        params = calib_data.get("params", {})
        ws_calib.append(["Calibration Results"])
        ws_calib.append([])
        ws_calib.append(["--- Calibrated Parameters ---"])
        ws_calib.append(["R1 [degC/W]", params.get("R1")])
        ws_calib.append(["R2 [degC/W]", params.get("R2")])
        ws_calib.append(["h_nat [W/(m^2*K)]", params.get("h_nat")])
        ws_calib.append(["h_rpm [W/(m^2*K)/sqrt(RPM)]", params.get("h_rpm")])
        ws_calib.append([])
        ws_calib.append(["--- Metrics ---"])
        ws_calib.append(["RMSE [degC]", calib_data.get("rmse")])
        ws_calib.append(["R-squared", calib_data.get("r_squared")])
        ws_calib.append(["Converged", calib_data.get("converged")])
        ws_calib.append(["Wall-clock time [s]", calib_data.get("time_s")])
        ws_calib.append([])
        ws_calib.append(["--- Loss History ---"])
        loss_hist = calib_data.get("loss_history", [])
        for i, val in enumerate(loss_hist):
            ws_calib.append([f"Start {i + 1}", val])

        # ---- Sheet 3: Simulated Temperatures ----
        T_coil_sim = calib_data.get("T_coil_sim", [])
        T_core_sim = calib_data.get("T_core_sim", [])
        T_housing_sim = calib_data.get("T_housing_sim", [])
        residuals = calib_data.get("residuals", [])

        if T_coil_sim:
            ws_temps = wb.create_sheet("T_core T_housing Estimate")
            ws_temps.append(["Index", "T_coil_sim [degC]", "T_core_sim [degC]", "T_housing_sim [degC]", "Residual [degC]"])
            for i in range(len(T_coil_sim)):
                ws_temps.append([
                    i + 1,
                    T_coil_sim[i] if i < len(T_coil_sim) else "",
                    T_core_sim[i] if i < len(T_core_sim) else "",
                    T_housing_sim[i] if i < len(T_housing_sim) else "",
                    residuals[i] if i < len(residuals) else "",
                ])

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"mtm_v2_{profile.name.replace(' ', '_')}_{profile.id[:8]}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
