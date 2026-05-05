"""Tests for compressor data extraction pipeline (REQ-COMP-DATA-001).

Covers:
  1. Parsing a valid multi-sheet Excel file
  2. Sheet with missing required columns -> rejected with error message
  3. Unit normalization (kPa->Pa, MPa->Pa, K->degC, g/s->kg/s, kW->W)
  4. Flexible column name matching (case, whitespace, aliases)
  5. Non-numeric value handling
  6. Empty sheet handling
  7. Parse result structure validation
  8. Empty rows skipped silently
  9. Duplicate RPM entries warn but keep
  10. Optional columns missing -> NaN/None, not rejected

Run with:
    cd backend && python -m pytest tests/test_compressor_data.py -v
"""

from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pytest

from core.compressor_data import (
    ColumnValidationResult,
    CompressorDataPoint,
    CompressorParseResult,
    SheetParseResult,
    normalize_units,
    parse_compressor_excel,
    validate_sheet_columns,
)


# ---------------------------------------------------------------------------
# Helpers: synthetic Excel fixture generators
# ---------------------------------------------------------------------------

# Canonical column names (all required + some optional)
REQUIRED_COLS = ["RPM", "Ps", "Ts", "Pd", "Td", "Tm", "P_loss"]
OPTIONAL_COLS = [
    "mass_flow", "hs", "Pm", "hm", "hd",
    "P1", "P2", "CompPower", "Torque", "I_peak",
    "IronLoss", "MotorLoss(AF)", "deltaML",
]
ALL_COLS = REQUIRED_COLS + OPTIONAL_COLS


def _build_valid_row(
    rpm: float = 3000.0,
    ps: float = 101325.0,
    ts: float = 25.0,
    pd: float = 500000.0,
    td: float = 80.0,
    tm: float = 50.0,
    p_loss: float = 150.0,
    mass_flow: float = 0.05,
    comp_power: float = 2000.0,
    torque: float = 6.5,
    motor_loss: float = 80.0,
) -> dict[str, float]:
    """Build a single valid data row with all columns."""
    return {
        "RPM": rpm,
        "Ps": ps,
        "Ts": ts,
        "Pd": pd,
        "Td": td,
        "Tm": tm,
        "P_loss": p_loss,
        "mass_flow": mass_flow,
        "hs": 420000.0,
        "Pm": 200000.0,
        "hm": 440000.0,
        "hd": 460000.0,
        "P1": 1200.0,
        "P2": 300.0,
        "CompPower": comp_power,
        "Torque": torque,
        "I_peak": 10.0,
        "IronLoss": 50.0,
        "MotorLoss(AF)": motor_loss,
        "deltaML": 5.0,
    }


def _write_excel(
    path: Path,
    sheets: dict[str, list[list[str | float | None]]],
) -> Path:
    """Write a multi-sheet Excel file from column headers + rows.

    sheets: {sheet_name: [header_row, data_row_1, data_row_2, ...]}
    Each row is a list of values matching the header length.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)

    wb.save(str(path))
    return path


def _valid_sheet_rows(
    n_points: int = 5,
    **overrides: float,
) -> list[list[str | float | None]]:
    """Build rows for a valid sheet with all columns."""
    header = list(ALL_COLS)
    rows: list[list[str | float | None]] = [header]
    for i in range(n_points):
        data = _build_valid_row(rpm=1000.0 * (i + 1), **overrides)
        rows.append([data.get(col) for col in ALL_COLS])
    return rows


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def tmp_excel(tmp_path: Path) -> Path:
    """Return path to a valid multi-sheet Excel file."""
    rows1 = _valid_sheet_rows(n_points=5)
    rows2 = _valid_sheet_rows(n_points=3)
    return _write_excel(
        tmp_path / "compressor_test.xlsx",
        {
            "Variant_A": rows1,
            "Variant_B": rows2,
        },
    )


@pytest.fixture()
def tmp_excel_kpa(tmp_path: Path) -> Path:
    """Excel file with kPa pressures (should be converted to Pa)."""
    header = ["RPM", "Ps[kPa]", "Ts", "Pd[kPa]", "Td", "Tm", "P_loss"]
    rows: list[list[str | float | None]] = [header]
    for i in range(3):
        rows.append([1000.0 * (i + 1), 101.325, 25.0, 500.0, 80.0, 50.0, 150.0])
    return _write_excel(tmp_path / "kpa_test.xlsx", {"Sheet1": rows})


@pytest.fixture()
def tmp_excel_mpa(tmp_path: Path) -> Path:
    """Excel file with MPa pressures."""
    header = ["RPM", "Ps[MPa]", "Ts", "Pd[MPa]", "Td", "Tm", "P_loss"]
    rows: list[list[str | float | None]] = [header]
    for i in range(3):
        rows.append([1000.0 * (i + 1), 0.101325, 25.0, 0.5, 80.0, 50.0, 150.0])
    return _write_excel(tmp_path / "mpa_test.xlsx", {"Sheet1": rows})


@pytest.fixture()
def tmp_excel_kelvin(tmp_path: Path) -> Path:
    """Excel file with Kelvin temperatures."""
    header = ["RPM", "Ps", "Ts[K]", "Pd", "Td[K]", "Tm[K]", "P_loss"]
    rows: list[list[str | float | None]] = [header]
    for i in range(3):
        rows.append([1000.0 * (i + 1), 101325.0, 298.15, 500000.0, 353.15, 323.15, 150.0])
    return _write_excel(tmp_path / "kelvin_test.xlsx", {"Sheet1": rows})


@pytest.fixture()
def tmp_excel_gs(tmp_path: Path) -> Path:
    """Excel file with g/s mass flow."""
    header = ["RPM", "Ps", "Ts", "Pd", "Td", "Tm", "P_loss", "mass_flow[g/s]"]
    rows: list[list[str | float | None]] = [header]
    for i in range(3):
        rows.append([1000.0 * (i + 1), 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 50.0])
    return _write_excel(tmp_path / "gs_test.xlsx", {"Sheet1": rows})


@pytest.fixture()
def tmp_excel_kw(tmp_path: Path) -> Path:
    """Excel file with kW power values."""
    header = ["RPM", "Ps", "Ts", "Pd", "Td", "Tm", "P_loss[kW]", "CompPower[kW]"]
    rows: list[list[str | float | None]] = [header]
    for i in range(3):
        rows.append([1000.0 * (i + 1), 101325.0, 25.0, 500000.0, 80.0, 50.0, 0.15, 2.0])
    return _write_excel(tmp_path / "kw_test.xlsx", {"Sheet1": rows})


# ===========================================================================
# Tests
# ===========================================================================


class TestColumnValidation:
    """Tests for validate_sheet_columns."""

    def test_all_required_present(self):
        result = validate_sheet_columns(list(ALL_COLS))
        assert isinstance(result, ColumnValidationResult)
        assert result.is_valid is True
        assert len(result.missing_required) == 0

    def test_missing_required_columns(self):
        cols = ["RPM", "Ts", "Td", "mass_flow"]  # missing Ps, Pd, Tm, P_loss
        result = validate_sheet_columns(cols)
        assert result.is_valid is False
        # Must report the missing required columns
        missing_lower = {c.lower() for c in result.missing_required}
        assert "ps" in missing_lower
        assert "pd" in missing_lower
        assert "tm" in missing_lower
        assert "p_loss" in missing_lower

    def test_optional_missing_does_not_fail(self):
        result = validate_sheet_columns(REQUIRED_COLS)
        assert result.is_valid is True
        # Optional columns should be listed as missing but not cause failure
        assert len(result.missing_optional) > 0

    def test_case_insensitive_matching(self):
        cols = ["rpm", "ps", "ts", "pd", "td", "tm", "p_loss"]
        result = validate_sheet_columns(cols)
        assert result.is_valid is True

    def test_whitespace_stripped(self):
        cols = [" RPM ", " Ps ", " Ts ", " Pd ", " Td ", " Tm ", " P_loss "]
        result = validate_sheet_columns(cols)
        assert result.is_valid is True

    def test_alias_matching(self):
        """Common aliases like MotorLoss, motor_loss, Motor Loss should match."""
        cols = REQUIRED_COLS + ["MotorLoss"]
        result = validate_sheet_columns(cols)
        # MotorLoss should match "MotorLoss(AF)" alias
        assert result.is_valid is True


class TestParseValidMultiSheet:
    """Tests for parsing a valid multi-sheet Excel file."""

    def test_parse_returns_result(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        assert isinstance(result, CompressorParseResult)

    def test_filename_extracted(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        assert result.filename == "compressor_test.xlsx"

    def test_sheet_count(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        assert result.valid_sheets == 2
        assert result.invalid_sheets == 0
        assert len(result.sheets) == 2

    def test_total_points(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        assert result.total_points == 8  # 5 + 3

    def test_per_sheet_metadata(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        sheet_names = {s.sheet_name for s in result.sheets}
        assert "Variant_A" in sheet_names
        assert "Variant_B" in sheet_names

        for sheet in result.sheets:
            assert isinstance(sheet, SheetParseResult)
            assert sheet.n_points > 0
            assert len(sheet.columns_found) >= len(REQUIRED_COLS)
            assert len(sheet.columns_missing) == 0
            assert len(sheet.errors) == 0
            assert len(sheet.data) == sheet.n_points

    def test_data_point_fields(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        sheet = result.sheets[0]
        dp = sheet.data[0]
        assert isinstance(dp, CompressorDataPoint)
        # Required fields must be populated
        assert dp.RPM > 0
        assert dp.Ps > 0
        assert dp.Ts > 0
        assert dp.Pd > 0
        assert dp.Td > 0
        assert dp.Tm > 0
        assert dp.P_loss > 0
        # Optional fields should also be present (our fixture has all cols)
        assert dp.mass_flow > 0
        assert dp.CompPower > 0
        assert dp.Torque > 0


class TestMissingRequiredColumns:
    """Sheets missing required columns should be rejected with error."""

    def test_missing_required_rejects_sheet(self, tmp_path: Path):
        header = ["RPM", "Ts", "Td", "mass_flow"]  # missing Ps, Pd, Tm, P_loss
        rows = [header, [3000.0, 25.0, 80.0, 0.05]]
        path = _write_excel(tmp_path / "missing.xlsx", {"BadSheet": rows})

        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 0
        assert result.invalid_sheets == 1

        sheet = result.sheets[0]
        assert len(sheet.errors) > 0
        assert sheet.n_points == 0
        assert len(sheet.data) == 0
        # Error should mention the missing columns
        error_text = " ".join(sheet.errors).lower()
        assert "ps" in error_text or "p_loss" in error_text

    def test_mixed_valid_invalid_sheets(self, tmp_path: Path):
        """Valid sheets proceed, invalid sheets rejected in same file."""
        valid_rows = _valid_sheet_rows(n_points=2)
        bad_header = ["RPM", "Ts", "Td"]
        bad_rows = [bad_header, [3000.0, 25.0, 80.0]]

        path = _write_excel(
            tmp_path / "mixed.xlsx",
            {"GoodSheet": valid_rows, "BadSheet": bad_rows},
        )
        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 1
        assert result.invalid_sheets == 1
        assert result.total_points == 2


class TestUnitNormalization:
    """Tests for normalize_units and automatic unit detection."""

    def test_kpa_to_pa(self, tmp_excel_kpa: Path):
        result = parse_compressor_excel(str(tmp_excel_kpa))
        sheet = result.sheets[0]
        assert result.valid_sheets == 1
        # Ps was 101.325 kPa -> should be ~101325 Pa
        dp = sheet.data[0]
        assert abs(dp.Ps - 101325.0) < 1.0

    def test_mpa_to_pa(self, tmp_excel_mpa: Path):
        result = parse_compressor_excel(str(tmp_excel_mpa))
        sheet = result.sheets[0]
        dp = sheet.data[0]
        # Ps was 0.101325 MPa -> ~101325 Pa
        assert abs(dp.Ps - 101325.0) < 1.0
        # Pd was 0.5 MPa -> 500000 Pa
        assert abs(dp.Pd - 500000.0) < 1.0

    def test_kelvin_to_celsius(self, tmp_excel_kelvin: Path):
        result = parse_compressor_excel(str(tmp_excel_kelvin))
        sheet = result.sheets[0]
        dp = sheet.data[0]
        # Ts was 298.15 K -> 25 degC
        assert abs(dp.Ts - 25.0) < 0.01
        # Td was 353.15 K -> 80 degC
        assert abs(dp.Td - 80.0) < 0.01
        # Tm was 323.15 K -> 50 degC
        assert abs(dp.Tm - 50.0) < 0.01

    def test_gs_to_kg_s(self, tmp_excel_gs: Path):
        result = parse_compressor_excel(str(tmp_excel_gs))
        sheet = result.sheets[0]
        dp = sheet.data[0]
        # mass_flow was 50 g/s -> 0.05 kg/s
        assert abs(dp.mass_flow - 0.05) < 0.001

    def test_kw_to_w(self, tmp_excel_kw: Path):
        result = parse_compressor_excel(str(tmp_excel_kw))
        sheet = result.sheets[0]
        dp = sheet.data[0]
        # P_loss was 0.15 kW -> 150 W
        assert abs(dp.P_loss - 150.0) < 0.1
        # CompPower was 2.0 kW -> 2000 W
        assert abs(dp.CompPower - 2000.0) < 0.1

    def test_normalize_units_function_directly(self):
        """Test normalize_units with a DataFrame directly."""
        import pandas as pd

        df = pd.DataFrame({
            "Ps[kPa]": [100.0, 200.0],
            "Pd[MPa]": [0.5, 1.0],
            "Ts[K]": [300.0, 310.0],
            "Td": [80.0, 90.0],
        })
        result = normalize_units(df)
        # Ps should be in Pa now
        assert abs(result.iloc[0]["Ps[kPa]"] - 100000.0) < 1.0
        # Pd should be in Pa
        assert abs(result.iloc[0]["Pd[MPa]"] - 500000.0) < 1.0
        # Ts should be in degC
        assert abs(result.iloc[0]["Ts[K]"] - 26.85) < 0.1
        # Td was already degC, unchanged
        assert abs(result.iloc[0]["Td"] - 80.0) < 0.01

    def test_kg_h_to_kg_s(self, tmp_path: Path):
        """mass_flow in kg/h -> kg/s."""
        header = ["RPM", "Ps", "Ts", "Pd", "Td", "Tm", "P_loss", "mass_flow[kg/h]"]
        rows: list[list[str | float | None]] = [header]
        for i in range(3):
            rows.append([1000.0 * (i + 1), 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 180.0])
        path = _write_excel(tmp_path / "kgh_test.xlsx", {"Sheet1": rows})

        result = parse_compressor_excel(str(path))
        dp = result.sheets[0].data[0]
        # 180 kg/h -> 0.05 kg/s
        assert abs(dp.mass_flow - 0.05) < 0.001


class TestFlexibleColumnNameMatching:
    """Column names should match flexibly (case, whitespace, aliases)."""

    def test_mixed_case_columns(self, tmp_path: Path):
        header = ["rpm", "PS", "ts", "pd", "TD", "tm", "p_loss", "Mass_Flow", "CompPower", "Torque", "MotorLoss"]
        rows: list[list[str | float | None]] = [header]
        for i in range(2):
            rows.append([1000.0 * (i + 1), 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 0.05, 2000.0, 6.5, 80.0])
        path = _write_excel(tmp_path / "mixed_case.xlsx", {"Sheet1": rows})

        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 1
        assert len(result.sheets[0].errors) == 0

    def test_whitespace_in_columns(self, tmp_path: Path):
        header = [" RPM ", "Ps ", " Ts", "Pd", " Td ", "Tm", "P_loss", "Motor Loss", "Motor Loss(AF)"]
        rows: list[list[str | float | None]] = [header]
        for i in range(2):
            rows.append([3000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 80.0, 80.0])
        path = _write_excel(tmp_path / "whitespace.xlsx", {"Sheet1": rows})

        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 1

    def test_underscore_alias(self, tmp_path: Path):
        """motor_loss should match MotorLoss(AF)."""
        header = ["RPM", "Ps", "Ts", "Pd", "Td", "Tm", "P_loss", "motor_loss"]
        rows: list[list[str | float | None]] = [header]
        for i in range(2):
            rows.append([3000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 80.0])
        path = _write_excel(tmp_path / "underscore.xlsx", {"Sheet1": rows})

        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 1
        # MotorLoss should be recognized as an optional column
        dp = result.sheets[0].data[0]
        assert abs(dp.MotorLoss - 80.0) < 0.01


class TestNonNumericValues:
    """Non-numeric cell values should cause row-level rejection."""

    def test_non_numeric_in_required_column(self, tmp_path: Path):
        header = list(REQUIRED_COLS) + ["CompPower", "Torque"]
        rows: list[list[str | float | None]] = [header]
        # Valid row
        rows.append([3000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])
        # Invalid row: RPM is a string
        rows.append(["invalid", 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])
        # Another valid row
        rows.append([4000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])

        path = _write_excel(tmp_path / "non_numeric.xlsx", {"Sheet1": rows})
        result = parse_compressor_excel(str(path))

        sheet = result.sheets[0]
        assert sheet.n_points == 2  # Only 2 valid rows
        assert len(sheet.errors) > 0  # Should report the bad row

    def test_non_numeric_in_optional_column(self, tmp_path: Path):
        """Non-numeric in optional column: row still parsed, optional field None."""
        header = list(ALL_COLS)
        rows: list[list[str | float | None]] = [header]
        # Row with invalid mass_flow
        vals = _build_valid_row()
        row_vals = [vals.get(col) for col in ALL_COLS]
        mass_flow_idx = ALL_COLS.index("mass_flow")
        row_vals[mass_flow_idx] = "N/A"
        rows.append(row_vals)

        path = _write_excel(tmp_path / "bad_optional.xlsx", {"Sheet1": rows})
        result = parse_compressor_excel(str(path))
        sheet = result.sheets[0]
        assert sheet.n_points == 1
        dp = sheet.data[0]
        # mass_flow should be NaN since it was invalid
        assert dp.mass_flow != dp.mass_flow  # NaN check


class TestEmptySheet:
    """Empty sheets should be handled gracefully."""

    def test_empty_sheet_rejected(self, tmp_path: Path):
        """Sheet with only a header (no data) should be rejected."""
        header = list(REQUIRED_COLS)
        path = _write_excel(tmp_path / "empty.xlsx", {"EmptySheet": [header]})
        result = parse_compressor_excel(str(path))
        assert result.invalid_sheets == 1
        sheet = result.sheets[0]
        assert sheet.n_points == 0

    def test_completely_empty_sheet(self, tmp_path: Path):
        """Sheet with no content at all."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet(title="BlankSheet")
        path = tmp_path / "blank.xlsx"
        wb.save(str(path))

        result = parse_compressor_excel(str(path))
        assert result.invalid_sheets == 1


class TestEmptyRows:
    """Empty rows should be skipped silently."""

    def test_empty_rows_skipped(self, tmp_path: Path):
        header = list(REQUIRED_COLS) + ["CompPower", "Torque"]
        rows: list[list[str | float | None]] = [header]
        # Valid row
        rows.append([3000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])
        # Empty row (all None)
        rows.append([None] * len(header))
        # Another valid row
        rows.append([4000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])

        path = _write_excel(tmp_path / "empty_rows.xlsx", {"Sheet1": rows})
        result = parse_compressor_excel(str(path))
        sheet = result.sheets[0]
        assert sheet.n_points == 2
        assert len(sheet.errors) == 0  # Empty rows silently skipped


class TestDuplicateRPM:
    """Duplicate RPM entries should warn but keep all rows."""

    def test_duplicate_rpm_warns(self, tmp_path: Path):
        header = list(REQUIRED_COLS) + ["CompPower", "Torque"]
        rows: list[list[str | float | None]] = [header]
        # Two rows with same RPM
        rows.append([3000.0, 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0, 2000.0, 6.5])
        rows.append([3000.0, 101325.0, 25.0, 600000.0, 90.0, 55.0, 160.0, 2100.0, 7.0])

        path = _write_excel(tmp_path / "dup_rpm.xlsx", {"Sheet1": rows})
        result = parse_compressor_excel(str(path))
        sheet = result.sheets[0]
        assert sheet.n_points == 2  # Both kept
        # Should have a warning about duplicates
        warning_text = " ".join(sheet.errors).lower()
        assert "duplicate" in warning_text or "rpm" in warning_text


class TestOptionalColumnsMissing:
    """Missing optional columns should use NaN/None, not reject."""

    def test_only_required_columns(self, tmp_path: Path):
        header = list(REQUIRED_COLS)
        rows: list[list[str | float | None]] = [header]
        for i in range(3):
            rows.append([1000.0 * (i + 1), 101325.0, 25.0, 500000.0, 80.0, 50.0, 150.0])

        path = _write_excel(tmp_path / "req_only.xlsx", {"Sheet1": rows})
        result = parse_compressor_excel(str(path))
        assert result.valid_sheets == 1
        sheet = result.sheets[0]

        # Optional fields should be NaN
        dp = sheet.data[0]
        assert math.isnan(dp.mass_flow)
        assert math.isnan(dp.CompPower)
        assert math.isnan(dp.Torque)

        # But missing_optional should be reported
        assert len(sheet.columns_missing) > 0


class TestParseResultStructure:
    """Validate the structure of parse results."""

    def test_compressor_parse_result_fields(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        assert hasattr(result, "filename")
        assert hasattr(result, "sheets")
        assert hasattr(result, "total_points")
        assert hasattr(result, "valid_sheets")
        assert hasattr(result, "invalid_sheets")

    def test_sheet_parse_result_fields(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        sheet = result.sheets[0]
        assert hasattr(sheet, "sheet_name")
        assert hasattr(sheet, "variant_name")
        assert hasattr(sheet, "n_points")
        assert hasattr(sheet, "columns_found")
        assert hasattr(sheet, "columns_missing")
        assert hasattr(sheet, "data")
        assert hasattr(sheet, "errors")

    def test_data_point_structure(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        dp = result.sheets[0].data[0]
        expected_fields = [
            "RPM", "mass_flow", "Ps", "Ts", "hs", "Pm", "Tm",
            "hm", "Pd", "Td", "hd", "P_loss", "P1", "P2",
            "CompPower", "Torque", "I_peak", "IronLoss", "MotorLoss", "deltaML",
        ]
        for field in expected_fields:
            assert hasattr(dp, field), f"CompressorDataPoint missing field: {field}"

    def test_variant_name_from_sheet_name(self, tmp_excel: Path):
        result = parse_compressor_excel(str(tmp_excel))
        variant_names = {s.variant_name for s in result.sheets}
        assert "Variant_A" in variant_names
        assert "Variant_B" in variant_names

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            parse_compressor_excel("/nonexistent/path/file.xlsx")
